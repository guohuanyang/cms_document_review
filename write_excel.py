# -*- coding: utf-8 -*-

# -------------------------------------------------------------------------------
# Name:         write_excel
# Description:
# Author:       guohuanyang
# Date:         2021/3/10
# Email         guohuanyang@datagrand.com
# -------------------------------------------------------------------------------
import os
from datetime import datetime

import openpyxl
from config import dir_name
sheet = []


class WriteExcel:
    """
    按照模版写入excel文件
    """

    def __init__(self, header=None):
        wb = openpyxl.Workbook()
        self.wb = wb
        self.ws = wb.active
        self.line_num = 1
        self.merge_level = 2
        self.header = header or list()

    def write_by_xy(self, data):
        start = 0
        end = data['level'] * self.merge_level
        if end > 1:
            start = end - 1
            self.ws.cell(self.line_num + 1, start + 1, data['ins_name'])
            data['cell_1'] = {
                "row": self.line_num + 1,
                "col": start + 1,
            }
            self.ws.cell(self.line_num + 1, start + 2, "%s%%" % data['ratio'])
            data['cell_2'] = {
                "row": self.line_num + 1,
                "col": start + 2,
            }
            # self.ws.cell(self.line_num + 1, start + 3, data['msg'] or '正常')
        else:
            self.ws.cell(self.line_num + 1, start + 1, data['ins_name'])
            data['cell_1'] = {
                "row": self.line_num + 1,
                "col": start + 1,
            }

    def write_header(self):
        for index, level in enumerate(self.header, 1):
            if index == 1:
                self.ws.cell(1, index, level)
            else:
                col = (index - 1) * self.merge_level
                self.ws.cell(1, col, level)
                self.ws.merge_cells(start_row=1, end_row=1, start_column=col, end_column=col+1)

    def dfs_write(self, root):
        if not root:
            return
        data = root['data']
        self.write_by_xy(data)
        if len(root['children'])==0:
            self.line_num += 1
        for children in root['children']:
            self.dfs_write(children)
        self.merge_children(root)

    def merge_children(self, root):
        children = root['children']
        # 孩子节点数量大于1的
        if len(children) > 1:
            # 该节点的行
            start_row_num = root['data']['cell_1']['row']
            # 最后一个孩子节点的行
            end_row_num = children[-1]['data']['cell_1']['row']
            # 该节点的列
            start_col_num = root['data']['cell_1']['col']
            # 根节点
            if start_col_num == 1:
                self.ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num,
                                    end_column=start_col_num)
                root['data']['cell_1'] = {
                    "row": end_row_num,
                    "col": start_col_num,
                }
            else:
                self.ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num,
                                    end_column=start_col_num)
                root['data']['cell_1'] = {
                    "row": end_row_num,
                    "col": start_col_num,
                }
                self.ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num+1,
                                    end_column=start_col_num+1)
                root['data']['cell_2'] = {
                    "row": end_row_num,
                    "col": start_col_num+1,
                }
        elif len(children) == 1:
            start_row_num = root['data']['cell_1']['row']
            end_row_num = children[0]['data']['cell_1']['row']
            start_col_num = root['data']['cell_1']['col']
            if start_col_num == 1:
                self.ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num,
                                    end_column=start_col_num)
                root['data']['cell_1'] = {
                    'row': end_row_num,
                    'col': start_col_num
                }
            else:
                self.ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num,
                                    end_column=start_col_num)
                root['data']['cell_1'] = {
                    'row': end_row_num,
                    'col': start_col_num
                }
                self.ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num+1,
                                    end_column=start_col_num+1)
                root['data']['cell_2'] = {
                    'row': end_row_num,
                    'col': start_col_num+1
                }

    def save(self, save_dir=dir_name):
        file_name = datetime.now().strftime("%Y%m%d-%H%M%S") + '.xlsx'
        if not os.path.exists(dir_name):
            os.mkdir(dir_name)
        excel_file_path = "%s/%s" % (save_dir, file_name)
        self.wb.save(excel_file_path)
        self.wb.close()
